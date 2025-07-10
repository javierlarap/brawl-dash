import os
import time
import requests
from urllib.parse import quote
from collections import defaultdict
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Font

API_KEY = "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzUxMiIsImtpZCI6IjI4YTMxOGY3LTAwMDAtYTFlYi03ZmExLTJjNzQzM2M2Y2NhNSJ9.eyJpc3MiOiJzdXBlcmNlbGwiLCJhdWQiOiJzdXBlcmNlbGw6Z2FtZWFwaSIsImp0aSI6Ijk1MzcyYzU3LTZhODctNDBmYS1iYzAzLWM2YWJlNzUyYmIyOCIsImlhdCI6MTc1MjE4NTQyOCwic3ViIjoiZGV2ZWxvcGVyL2EwYjQ4NGMyLWMzMjAtYWY3Yi1lOTJjLTI1Y2JhOTM4YTExNCIsInNjb3BlcyI6WyJicmF3bHN0YXJzIl0sImxpbWl0cyI6W3sidGllciI6ImRldmVsb3Blci9zaWx2ZXIiLCJ0eXBlIjoidGhyb3R0bGluZyJ9LHsiY2lkcnMiOlsiMTguMTU2LjE1OC41MyIsIjE4LjE1Ni40Mi4yMDAiLCI1Mi41OS4xMDMuNTQiXSwidHlwZSI6ImNsaWVudCJ9XX0.-GPYUZKYi7g2zXfOpIQzkL3FFPt8yru2ieo5rMliFoJOC4bY8_81-oyJSLzJiCrYCtgsMDjfBmwRdvFNDcOfRw"
HEADERS = {"Authorization": f"Bearer {API_KEY}"}

PRO_TAGS = [
    "#P90RJQ8C", "#8Y98Q8U", "#29UGLJV2G", "#J99YU9QY", "#82PQUPGU0",
    "#28PU0P9L0", "#2UQCRP90", "#LJ0288PRG", "#8Y2Y0GYYG", "#2G0RRLU2R",
    "#2PV9R9Q28", "#P0Y8JGL0U", "#9ULYPV8", "#20C0LL00", "#UR2UL8YR",
    "#9JCG0VY8U", "#9LVUC2PY", "#CJV2PJ0R", "#9PCV9L982", "#2Q892QVU",
    "#82RGU8PR", "#2VJCCCQGP", "#22CL00PG0", "#90JCYPQU", "#2L08U08RUQ",
    "#P2808PRC", "#J2L2C0RJ", "#YQUCCJ2", "#V89Y2GP0", "#PCPRPJV",
    "#2JLVRQ9", "#20L88L2J", "#2YYJYPCL", "#9PQQ8GQQ", "#PLV89CGP",
    "#Q808R2CV", "#2288RRJJG", "#29VRJU08C", "#9RVPL0Q0P", "#R0VQ82GU",
    "#88Y8UGPG", "#Q22ULY9JY", "#2JQU8JQ2G", "#90CUVVL2R", "#80PVPCC29",
    "#QLVP829R", "#2RUUJJ8U9", "#U9GC8G02", "#82RCQCVG", "#2Q028GQQP",
    "#PR9U2JL", "#GCJCRVQ8", "#GVLRUG9Q", "#QLCJGQUP", "#2PGGR8Y9P",
    "#JJ09PC0P", "#R9CCLP8Q", "#Q2VCLG9Y9", "#QURVLPG", "#LYR0Q9C",
    "#2LG8QJ9L", "#VPVLG2", "#LVRRYPV", "#VJUQ0Y", "#QUYCVC2", "#8V92UYCJ", "#2LQ0RGCRU", "#2LL892UU", "#QUG9RP9", "#2RR2RU8UL","#2Y822YJYJC", "#Y8PLP8VY", "#2P9CJVGJ8","#9GPQR8CGL", "#8CV8PCGC","#99GGUPY2U", "#88LLQGP0Q", "#89UUQLJCC", "#80YVJGRY", "#8GUPLYY", "#2G2P99ULJP", "#CPPC8282", "#8CYJ8QGR", "#L08Q9J09","#GGUQCG0G","#QJULVGU","#JQ8LLLY","#JQ8L0YYL","#PLLRJC2V","#RVL0RPR9","#L9PQUV0YC","#9QCJPL20","#9CPYUCGQC","#YQLP8LP8","#2QG9LQQC8Y","#GVYLVUGR","#YGQYGCR","#2CJ0RCJ","#9JYG98GG","#R2LR2QLG","#JVRCVJ9Q","#202GJJR28","#89PVJG9R0","#PR0P8QVQ","#GYCYCLRJL","#820JCJJG","#RCYQUJU0","#GJPVYUQG"
]

MAP_WHITELIST = [
    "Hard Rock Mine", "Double Swoosh", "Crystal Arcade", "Ring of Fire",
    "Open Business", "Dueling Beetles", "Hot Potato", "Pit Stop",
    "Kaboom Canyon", "Triple Dribble", "Pinball Dreams", "Sneaky Fields",
    "Dry Season", "Hideout", "Layer Cake", "Goldarm Gulch", "Belle's Rock",
    "New Horizons"
]

def get_battlelog(tag):
    url = f"https://api.brawlstars.com/v1/players/{quote(tag)}/battlelog"
    r = requests.get(url, headers=HEADERS)
    return r.json().get("items", []) if r.status_code == 200 else []

def extract(team):
    brawlers = [p.get("brawler", {}).get("name", "(?)") for p in team]
    names = [p.get("name", "") for p in team]
    return brawlers, names

def load_existing_timestamps(filename):
    seen_ts = set()
    if not os.path.exists(filename):
        return seen_ts

    wb = load_workbook(filename, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=4, values_only=True):
            if len(row) < 14:
                continue
            ts = row[13]
            if ts:
                ts_clean = str(ts).strip().replace('\u200b', '').replace('\n', '').replace('\r', '')
                seen_ts.add(ts_clean)
    print(f"\n✅ Total de timestamps cargados: {len(seen_ts)}")
    return seen_ts

def detect_scrims_unicos(existing_ts):
    scrims_raw = []
    pro_set = set(PRO_TAGS)

    for tag in PRO_TAGS:
        for b in get_battlelog(tag):
            ts = b.get("battleTime")
            if not ts:
                continue

            info = b.get("battle", {})
            if info.get("type") != "friendly":
                continue

            map_name = b.get("event", {}).get("map", "")
            if map_name not in MAP_WHITELIST:
                continue

            teams = info.get("teams", [])
            if len(teams) != 2 or any(len(t) != 3 for t in teams):
                continue

            ptags = {p.get("tag", "").upper() for t in teams for p in t}
            if len(ptags & pro_set) < 3:
                continue

            b_e1, n_e1 = extract(teams[0])
            b_e2, n_e2 = extract(teams[1])
            res = info.get("result", "")

            if res == "victory":
                winner = "Equipo 1"
            elif res == "defeat":
                winner = "Equipo 2"
            elif res == "draw":
                winner = "Empate"
            else:
                winner = "Desconocido"

            scrims_raw.append({
                "battle_time": ts,
                "map_name": map_name,
                "brawlers_e1": b_e1,
                "brawlers_e2": b_e2,
                "names_e1": n_e1,
                "names_e2": n_e2,
                "winner": winner
            })

    # Eliminar duplicados por (mapa, timestamp)
    unique_scrims = {}
    for s in scrims_raw:
        key = (s["map_name"], s["battle_time"])
        if key not in unique_scrims:
            unique_scrims[key] = s

    # Comparar con los que ya están en el Excel
    scrims_by_map = defaultdict(list)
    for (map_name, ts), s in unique_scrims.items():
        if ts not in existing_ts:
            scrims_by_map[map_name].append(s)

    print(f"\n✅ Scrims únicas nuevas detectadas: {sum(len(v) for v in scrims_by_map.values())}")
    return scrims_by_map

def save_scrims(scrims_by_map, filename="scrims_actualizado.xlsx"):
    blue = PatternFill("solid", fgColor="CCE5FF")
    red = PatternFill("solid", fgColor="F4CCCC")
    gray = PatternFill("solid", fgColor="D9D9D9")
    bold = Font(bold=True)
    thin = Border(*([Side("thin")] * 4))

    wb = load_workbook(filename) if os.path.exists(filename) else Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    for map_name, entries in scrims_by_map.items():
        ws = wb[map_name] if map_name in wb.sheetnames else wb.create_sheet(map_name[:31])

        if ws.max_row < 3:
            headers = (
                ["B1", "B2", "B3", "B1", "B2", "B3", "Ganador"] +
                [f"Jugador {i}" for i in range(1, 7)] +
                ["Timestamp"]
            )
            for col, h in enumerate(headers, start=1):
                c = ws.cell(row=3, column=col, value=h)
                c.border = thin
                if col <= 3:
                    c.fill = blue
                elif 4 <= col <= 6:
                    c.fill = red

        ts_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=3, column=col).value == "Timestamp":
                ts_col = col
                break
        if not ts_col:
            ts_col = ws.max_column + 1

        start_row = ws.max_row + 1
        while any(ws.cell(row=start_row, column=col).value for col in range(1, ts_col + 1)):
            start_row += 1

        print(f"\n📥 Guardando {len(entries)} scrims en hoja '{map_name}' desde fila {start_row}")

        for i, e in enumerate(entries, start=start_row):
            for j, b in enumerate(e["brawlers_e1"], start=1):
                c = ws.cell(row=i, column=j, value=b)
                c.fill = blue
                c.border = thin
            for j, b in enumerate(e["brawlers_e2"], start=4):
                c = ws.cell(row=i, column=j, value=b)
                c.fill = red
                c.border = thin

            wc = ws.cell(row=i, column=7, value=e["winner"])
            wc.fill = {"Equipo 1": blue, "Equipo 2": red}.get(e["winner"], gray)
            wc.font = bold
            wc.border = thin

            for k, name in enumerate(e["names_e1"] + e["names_e2"], start=8):
                ws.cell(row=i, column=k, value=name).border = thin

            ts_cell = ws.cell(row=i, column=ts_col, value=e["battle_time"])
            ts_cell.border = thin

    wb.save(filename)
    print(f"\n✅ Excel actualizado con nuevas scrims: {filename}")

from threading import Thread
from flask import Flask

app = Flask(__name__)

@app.route("/")
def status():
    return "✅ Servicio activo y ejecutando scraping en segundo plano."

def iniciar_flask():
    app.run(host="0.0.0.0", port=10000)

# ────── BLOQUE PRINCIPAL ──────
if __name__ == "__main__":
    Thread(target=iniciar_flask).start()
    
    while True:
        print(f"\n⏱️ Iniciando ejecución a las {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        try:
            archivo = "scrims_actualizado.xlsx"

            print("📂 Cargando timestamps existentes...")
            existing_ts = load_existing_timestamps(archivo)

            print("🔍 Buscando scrims nuevas...")
            nuevas = detect_scrims_unicos(existing_ts)

            print("📊 Resultado:")
            print(f" - Timestamps cargados: {len(existing_ts)}")
            print(f" - Scrims nuevas detectadas: {sum(len(v) for v in nuevas.values())}")

            if nuevas:
                print("\n💾 Guardando nuevas scrims en Excel...")
                save_scrims(nuevas, archivo)
            else:
                print("\n⚠️ No hay scrims nuevas (solo timestamp).")

        except Exception as e:
            print(f"\n❌ Error durante la ejecución: {e}")

        print("🕒 Esperando 15 minutos...\n")
        time.sleep(900)  # 900 segundos = 15 minutos

